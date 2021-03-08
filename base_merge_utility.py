"""
ПОДГОТОВКА ФАЙЛОВ
1. все файлы баз и прайсов должны быть закрыты!
2. файл главной базы должен соответствовать маске [*_ad???@?adis???rg_*.xlsx]
3. файлы новых прайсов
    - содержать имя поставщика например "surgaz" (в любой раскладке)
    - файл прайс от одного вендора должен быть только 1!
"""
import openpyxl
import sys
import glob

# ================================================================
# 0=SETTINGS
find_header_line_square_int = 20

file_opened_startwith_symbols = "~$"

file_base_mask = "*_ad???@?adis???rg_*.xlsx"
base_header_detect_cell_value = "Группы"
base_column_detect_code = "Код"
base_column_detect_article = "Артикул"
base_column_detect_price3 = "Цена: Цена продажи"
base_column_detect_price2 = "Цена: РРЦ"
base_column_detect_price1 = "Закупочная цена"
base_column_values_code_null_set = {base_column_detect_code, None, ""}

vendor_dict = {"surgaz": {"file_mask": "*surgaz*.xlsx",
                          "file_found_if_one": None,

                          "header_detect_cell_value": "Артикул",
                          "column_detect_article": "Артикул",
                          "column_detect_price1": "Цена за рулон кратно рулону (Продажи1)",
                          "column_detect_price2": "МРЦ",
                          "column_detect_price3": None,     # need formula!!!!

                          "article_blank_set": {None, ""},
                          "article_mask_blank_set": {"Артикул", "Материал", "жизни", },

                          "data_article_all_dict": dict(),      # большой!
                          "data_article_repeated_set": set()},
               }

# ================================================================
# 1=DETECT FILES
# ----------------------------------------------------------------
# 1=FILE - MAIN BASE
print("-"*80)
print(f"START finding files MainBASE")

file_base_found_list = glob.glob(file_base_mask)
print("Найдены файлы главной базы:", file_base_found_list)


def files_found_several(type_txt):
    print(f"ОШИБКА: найдено несколько файлов {type_txt}! уберите лишние! и перезапустите приложение", file=sys.stderr)
    input("Нажмите Enter для выхода")
    sys.exit()


def files_found_opened(type_txt):
    print(f"ОШИБКА: найден открытый файл {type_txt}! закройте его и перезапустите приложение", file=sys.stderr)
    input("Нажмите Enter для выхода")
    sys.exit()


if len(file_base_found_list) == 1:
    file_base = file_base_found_list[0]
elif len(file_base_found_list) > 1:
    for file in file_base_found_list:
        if file.startswith(file_opened_startwith_symbols):
            files_found_opened("главной базы")
    else:
        files_found_several("главной базы")

# ----------------------------------------------------------------
# 2=FILES - VENDORS
print("-"*80)
print(f"START finding files VENDORS")

for vendor in vendor_dict:
    print("-" * 40)
    print(f"VENDOR [{vendor}]")

    vendor_data_dict = vendor_dict[vendor]
    file_vendor_mask = vendor_data_dict["file_mask"]
    file_vendor_found_list = glob.glob(file_vendor_mask)
    print(f"Найдены файлы вендора [{vendor}]: {file_vendor_found_list}")

    if len(file_vendor_found_list) == 1:
        vendor_data_dict["file_found_if_one"] = file_vendor_found_list[0]
    elif len(file_vendor_found_list) > 1:
        for file in file_vendor_found_list:
            if file.startswith(file_opened_startwith_symbols):
                files_found_opened(f"вендора [{vendor}]")
        else:
            files_found_several(f"вендора [{vendor}]")

# ================================================================
# 2=MainBASE=WORK
print("*"*80)
print(f"START MainBASE WORK")

# --------------------------
# 1=file INIT
wb_base = openpyxl.load_workbook(file_base)
ws_base = wb_base.active

# --------------------------
# 2=DETECT HEADER LINE ROW
cell_iter_cols = ws_base.iter_cols(max_row=find_header_line_square_int, max_col=find_header_line_square_int)
for column_tuple in cell_iter_cols:
    for cell_obj in column_tuple:
        cell_value = cell_obj.value
        if cell_value == base_header_detect_cell_value:
            base_header_row_int = cell_obj.row

# --------------------------
# 3=DETECT COLUMNS IN HEADER
row_title = ws_base[base_header_row_int]
column_seek_index = 1
for cell in row_title:
    cell_value = cell.value
    if cell_value == base_column_detect_code:
        base_column_index_code = column_seek_index
        print(f"Номер колонки [{base_column_detect_code}] = [{base_column_index_code}]")
    elif cell_value == base_column_detect_article:
        base_column_index_art = column_seek_index
        print(f"Номер колонки [{base_column_detect_article}] = [{base_column_index_art}]")
    elif cell_value == base_column_detect_price1:
        base_column_index_price1 = column_seek_index
        print(f"Номер колонки price1[{base_column_detect_price1}] = [{base_column_index_price1}]")
    elif cell_value == base_column_detect_price2:
        base_column_index_price2 = column_seek_index
        print(f"Номер колонки price2[{base_column_detect_price2}] = [{base_column_index_price2}]")
    elif cell_value == base_column_detect_price3:
        base_column_index_price3 = column_seek_index
        print(f"Номер колонки price3[{base_column_detect_price3}] = [{base_column_index_price3}]")
    column_seek_index += 1

# --------------------------
# 4=load DATA - ColumnCODE
column_values_code_base_all_dict = dict()
column_values_code_base_repeated_set = set()

print("-"*80)
print(f"load data from file: [{file_base}]")
print("-"*40)
column_values_code_base_iter = ws_base.iter_cols(min_col=base_column_index_code, max_col=base_column_index_code)
for column_tuple in column_values_code_base_iter:
    for cell_obj in column_tuple:
        cell_value = cell_obj.value
        if cell_value not in column_values_code_base_all_dict:
            # print("+", cell_value)
            column_values_code_base_all_dict.update({cell_value: {"cell_obj_list": [cell_obj, ],
                                                                  "price1": None,
                                                                  "price2": None,
                                                                  "price3": None,
                                                                  }})
            column_values_code_base_all_dict["price1"] = ws_base.cell(row=cell_obj.row, column=base_column_index_price1).value
            column_values_code_base_all_dict["price2"] = ws_base.cell(row=cell_obj.row, column=base_column_index_price2).value
            column_values_code_base_all_dict["price3"] = ws_base.cell(row=cell_obj.row, column=base_column_index_price3).value

        else:
            print(f'found repeated value: [{cell_value}]')
            column_values_code_base_all_dict[cell_value]["cell_obj_list"].append(cell_obj)
            column_values_code_base_repeated_set.update({cell_value})

# --------------------------
# 5=print loadRESULTS
count_column_values_code_base_all_dict = len(column_values_code_base_all_dict)
count_column_values_code_base_repeated_set = len(column_values_code_base_repeated_set)

print("-"*40)
# print("from file:", file_base)
print("count_column_values_code_base_all_dict:", count_column_values_code_base_all_dict)
print("count_column_values_code_base_repeated_set:", count_column_values_code_base_repeated_set)

print("column_values_code_base_repeated_set:", column_values_code_base_repeated_set)
print("-"*80)

# ================================================================
# 3=VENDOR FILES=WORK
print("*"*80)
print(f"START VENDOR WORK")

for vendor in vendor_dict:
    vendor_data_dict = vendor_dict[vendor]

    file_name_vendor = vendor_data_dict["file_found_if_one"]
    if file_name_vendor is None:
        continue

    # --------------------------
    # 1=file INIT
    wb_vendor = openpyxl.load_workbook(file_name_vendor)
    ws_vendor = wb_vendor.active

    # --------------------------
    # 2=DETECT HEADER LINE ROW
    cell_iter_cols = ws_vendor.iter_cols(max_row=find_header_line_square_int, max_col=find_header_line_square_int)
    for column_tuple in cell_iter_cols:
        for cell_obj in column_tuple:
            cell_value = cell_obj.value
            if cell_value == vendor_data_dict["header_detect_cell_value"]:
                vendor_header_row_int = cell_obj.row

    # --------------------------
    # 3=DETECT COLUMNS IN HEADER
    row_title = ws_vendor[vendor_header_row_int]
    column_seek_index = 1
    for cell in row_title:
        cell_value = cell.value
        if cell_value == vendor_data_dict["column_detect_article"]:
            vendor_column_index_article = column_seek_index
            print(f"Номер колонки [column_detect_article] = [{vendor_column_index_article}]")
        elif cell_value == vendor_data_dict["column_detect_price1"]:
            vendor_column_index_price1 = column_seek_index
            print(f"Номер колонки price1[vendor_column_detect_price1] = [{vendor_column_index_price1}]")
        elif cell_value == vendor_data_dict["column_detect_price2"]:
            vendor_column_index_price2 = column_seek_index
            print(f"Номер колонки price2[vendor_column_detect_price2] = [{vendor_column_index_price2}]")
        elif cell_value == vendor_data_dict["column_detect_price3"]:
            pass
            # vendor_column_index_price3 = column_seek_index
            # print(f"Номер колонки price3[vendor_column_detect_price3] = [{vendor_column_index_price3}]")
        column_seek_index += 1

    # --------------------------
    # 3=load DATA - ColumnCODE
    column_values_vendor_article_all_dict = vendor_data_dict["data_article_all_dict"]
    column_values_vendor_article_repeated_set = vendor_data_dict["data_article_repeated_set"]

    print("-"*80)
    print(f"load data from file: [{file_name_vendor}]")
    print("-"*40)
    column_values_vendor_article_iter = ws_vendor.iter_cols(min_col=vendor_column_index_article, max_col=vendor_column_index_article)
    for column_tuple in column_values_vendor_article_iter:
        for cell_obj in column_tuple:
            cell_value = cell_obj.value

            if cell_value not in column_values_vendor_article_all_dict:
                # print("+", cell_value)
                column_values_vendor_article_all_dict.update({cell_value: {"cell_obj_list": [cell_obj, ],
                                                                           "marker": None,
                                                                           "price1": None,
                                                                           "price2": None,
                                                                           "price3": None,
                                                                           }})
                """
                MARKER

                -1=INCORRECT DATA=exists different price for one article
                0=NULL ARTICLE
                1=INFO LINE

                99=clear old price!
                100=OK=renew price!
                101=NEW PRODUCT!!!
                """
                cell_value_dict = column_values_vendor_article_all_dict[cell_value]
                cell_obj_price1 = ws_vendor.cell(row=cell_obj.row, column=vendor_column_index_price1).value
                cell_obj_price2 = ws_vendor.cell(row=cell_obj.row, column=vendor_column_index_price2).value

                # check FOR BLANK
                if cell_value in vendor_data_dict["article_blank_set"]:
                    cell_value_dict["marker"] = 0   # NULL ARTICLE
                    continue
                elif all([mask in cell_value for mask in vendor_data_dict["article_mask_blank_set"]]):
                    cell_value_dict["marker"] = 1   # INFO LINE
                    continue

                # check FOR PRICE
                if cell_obj_price1 is None:
                    if column_values_code_base_all_dict.get(cell_value, None) is None:
                        cell_value_dict["marker"] = 1   # INFO LINE
                    else:
                        cell_value_dict["marker"] = 99   # CLEAR

                else:
                    cell_value_dict["price1"] = cell_obj_price1
                    cell_value_dict["price2"] = cell_obj_price2

                    if column_values_code_base_all_dict.get(cell_value, None) is None:
                        cell_value_dict["marker"] = 101      # NEW PRODUCT!!!
                    else:
                        cell_value_dict["marker"] = 100     # OK

            else:
                cell_value_dict = column_values_vendor_article_all_dict[cell_value]
                cell_obj_list = cell_value_dict["cell_obj_list"]
                cell_obj_list.append(cell_obj)

                # check indent price! may be it was several articles but indent price! it is OK!
                cell_obj_price1_last = ws_vendor.cell(row=cell_obj_list[-1].row, column=vendor_column_index_price1).value
                cell_obj_price1_prev = ws_vendor.cell(row=cell_obj_list[-2].row, column=vendor_column_index_price1).value
                if cell_obj_price1_last != cell_obj_price1_prev:
                    column_values_vendor_article_repeated_set.update({cell_value})
                    count_repeated = len(column_values_vendor_article_all_dict[cell_value]["cell_obj_list"])
                    print(f'found repeated value: [{cell_value}] \tby [{count_repeated}]times')
                    cell_value_dict["marker"] = -1   # INCORRECT DATA=exists different price for one article

    # --------------------------
    # 4=print loadRESULTS
    count_column_values_vendor_article_all_dict = len(column_values_vendor_article_all_dict)
    count_column_values_vendor_article_repeated_set = len(column_values_vendor_article_repeated_set)

    print("-"*40)
    # print("from file:", file_base)
    print("count_column_values_vendor_article_all_dict:", count_column_values_vendor_article_all_dict)
    print("count_column_values_vendor_article_repeated_set:", count_column_values_vendor_article_repeated_set)

    print("column_values_vendor_article_repeated_set:", column_values_vendor_article_repeated_set)
    print("*"*80)

    result_marker_dict = dict()
    for cell_value in column_values_vendor_article_all_dict:
        data_dict = column_values_vendor_article_all_dict[cell_value]
        article_value = cell_value
        article_mark = data_dict["marker"]

        vendor_article_price1 = data_dict["price1"]
        vendor_article_price2 = data_dict["price2"]
        vendor_article_price3 = data_dict["price3"]

        if article_mark == 100:
            base_article_price1 = column_values_code_base_all_dict[article_value]["price1"]
            base_article_price2 = column_values_code_base_all_dict[article_value]["price2"]
            base_article_price3 = column_values_code_base_all_dict[article_value]["price3"]

            print(f"{article_mark}={base_article_price1}/{base_article_price2}/{base_article_price3}[{article_value}]{vendor_article_price1}/{vendor_article_price2}/{vendor_article_price3}")

        if result_marker_dict.get(article_mark, None) is None:
            result_marker_dict.update({article_mark: set()})
        result_marker_dict[article_mark].update({article_value})

    print("*"*80)
    print("MARKERS STATISTICS")
    for marker in result_marker_dict:
        print(f"marker[{marker}]=[{len(result_marker_dict[marker])}]count", )
    print("*"*80)
